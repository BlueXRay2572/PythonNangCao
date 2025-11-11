from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, send_file
from flask_sqlalchemy import SQLAlchemy
from flask_admin import Admin
from flask_admin.contrib.sqla import ModelView
from datetime import datetime
import pandas as pd
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
import plotly.express as px
import plotly.graph_objects as go
import json
import plotly

app = Flask(__name__)
app.config['SECRET_KEY'] = 'your-secret-key-here'
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///inventory.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db = SQLAlchemy(app)
admin = Admin(app, name='Inventory Admin')

# Models
class Warehouse(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    location = db.Column(db.String(200))
    products = db.relationship('Product', backref='warehouse', lazy=True)

class Supplier(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    contact = db.Column(db.String(100))
    email = db.Column(db.String(100))
    phone = db.Column(db.String(20))
    products = db.relationship('Product', backref='supplier', lazy=True)

class Category(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    products = db.relationship('Product', backref='category', lazy=True)

class Product(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    sku = db.Column(db.String(50), unique=True, nullable=False)
    barcode = db.Column(db.String(100), unique=True)
    name = db.Column(db.String(200), nullable=False)
    description = db.Column(db.Text)
    quantity = db.Column(db.Integer, default=0)
    min_stock = db.Column(db.Integer, default=10)
    price = db.Column(db.Float, default=0)
    category_id = db.Column(db.Integer, db.ForeignKey('category.id'))
    supplier_id = db.Column(db.Integer, db.ForeignKey('supplier.id'))
    warehouse_id = db.Column(db.Integer, db.ForeignKey('warehouse.id'))
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    transactions = db.relationship('Transaction', backref='product', lazy=True)
    batch_number = db.Column(db.String(100), nullable=True)
    expiry_date = db.Column(db.DateTime, nullable=True)

class Transaction(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    product_id = db.Column(db.Integer, db.ForeignKey('product.id'), nullable=False)
    transaction_type = db.Column(db.String(20), nullable=False)  # IN or OUT
    quantity = db.Column(db.Integer, nullable=False)
    notes = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

# Flask-Admin Views
admin.add_view(ModelView(Product, db.session))
admin.add_view(ModelView(Category, db.session))
admin.add_view(ModelView(Supplier, db.session))
admin.add_view(ModelView(Warehouse, db.session))
admin.add_view(ModelView(Transaction, db.session))

# Routes
@app.route('/')
def index():
    products = Product.query.all()
    low_stock = Product.query.filter(Product.quantity <= Product.min_stock).all()
    total_products = len(products)
    total_value = sum([p.quantity * p.price for p in products])
    warehouses = Warehouse.query.count()
    suppliers = Supplier.query.count()
    
    return render_template('index.html', 
                         products=products,
                         low_stock=low_stock,
                         total_products=total_products,
                         total_value=total_value,
                         warehouses=warehouses,
                         suppliers=suppliers)

@app.route('/products')
def products():
    products = Product.query.all()
    categories = Category.query.all()
    suppliers = Supplier.query.all()
    warehouses = Warehouse.query.all()
    return render_template('products.html', 
                         products=products,
                         categories=categories,
                         suppliers=suppliers,
                         warehouses=warehouses)

@app.route('/product/add', methods=['POST'])
def add_product():
    try:
        expiry_date_str = request.form.get('expiry_date')
        expiry_date = datetime.strptime(expiry_date_str, '%Y-%m-%d') if expiry_date_str else None
        product = Product(
            sku=request.form['sku'],
            barcode=request.form.get('barcode'),
            name=request.form['name'],
            description=request.form.get('description'),
            quantity=int(request.form.get('quantity', 0)),
            min_stock=int(request.form.get('min_stock', 10)),
            price=float(request.form.get('price', 0)),
            category_id=int(request.form['category_id']) if request.form.get('category_id') else None,
            supplier_id=int(request.form['supplier_id']) if request.form.get('supplier_id') else None,
            warehouse_id=int(request.form['warehouse_id']) if request.form.get('warehouse_id') else None,
            batch_number=request.form.get('batch_number'),
            expiry_date=expiry_date
        )
        db.session.add(product)
        db.session.commit()
        flash('Sản phẩm đã được thêm thành công!', 'success')
    except Exception as e:
        flash(f'Lỗi: {str(e)}', 'danger')
    return redirect(url_for('products'))

@app.route('/product/delete/<int:id>')
def delete_product(id):
    product = Product.query.get_or_404(id)
    db.session.delete(product)
    db.session.commit()
    flash('Sản phẩm đã được xóa!', 'success')
    return redirect(url_for('products'))

@app.route('/transaction', methods=['POST'])
def add_transaction():
    try:
        product_id = int(request.form['product_id'])
        trans_type = request.form['type']
        quantity = int(request.form['quantity'])
        notes = request.form.get('notes', '')
        
        product = Product.query.get_or_404(product_id)
        
        if trans_type == 'IN':
            product.quantity += quantity
        else:  # OUT
            if product.quantity < quantity:
                flash('Không đủ số lượng trong kho!', 'danger')
                return redirect(url_for('index'))
            product.quantity -= quantity
        
        transaction = Transaction(
            product_id=product_id,
            transaction_type=trans_type,
            quantity=quantity,
            notes=notes
        )
        
        db.session.add(transaction)
        db.session.commit()
        flash('Giao dịch đã được ghi nhận!', 'success')
    except Exception as e:
        flash(f'Lỗi: {str(e)}', 'danger')
    
    return redirect(url_for('index'))

@app.route('/reports')
def reports():
    products = Product.query.all()
    transactions = Transaction.query.order_by(Transaction.created_at.desc()).limit(100).all()
    
    # Tạo DataFrame cho phân tích
    product_data = [{
        'SKU': p.sku,
        'Tên sản phẩm': p.name,
        'Số lượng': p.quantity,
        'Giá': p.price,
        'Tổng giá trị': p.quantity * p.price,
        'Danh mục': p.category.name if p.category else 'N/A',
        'Nhà cung cấp': p.supplier.name if p.supplier else 'N/A',
        'Kho': p.warehouse.name if p.warehouse else 'N/A'
    } for p in products]
    
    df = pd.DataFrame(product_data)
    
    # Tạo biểu đồ
    charts = {}
    
    # Biểu đồ phân bố số lượng theo danh mục
    if not df.empty and 'Danh mục' in df.columns:
        category_chart = px.bar(df.groupby('Danh mục')['Số lượng'].sum().reset_index(), 
                               x='Danh mục', y='Số lượng',
                               title='Số lượng tồn kho theo danh mục')
        charts['category'] = json.dumps(category_chart, cls=plotly.utils.PlotlyJSONEncoder)
        
        # Biểu đồ giá trị theo kho
        warehouse_chart = px.pie(df.groupby('Kho')['Tổng giá trị'].sum().reset_index(),
                                values='Tổng giá trị', names='Kho',
                                title='Phân bố giá trị theo kho')
        charts['warehouse'] = json.dumps(warehouse_chart, cls=plotly.utils.PlotlyJSONEncoder)
    
    return render_template('reports.html', 
                         products=products,
                         transactions=transactions,
                         charts=charts,
                         df=df.to_html(classes='table table-striped') if not df.empty else '')

@app.route('/export/excel')
def export_excel():
    products = Product.query.all()
    
    # Tạo workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventory Report"
    
    # Header
    headers = ['SKU', 'Barcode', 'Tên sản phẩm', 'Số lượng', 'Mức tối thiểu', 
               'Giá', 'Tổng giá trị', 'Danh mục', 'Nhà cung cấp', 'Kho']
    ws.append(headers)
    
    # Style header
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    # Data
    for p in products:
        ws.append([
            p.sku,
            p.barcode or '',
            p.name,
            p.quantity,
            p.min_stock,
            p.price,
            p.quantity * p.price,
            p.category.name if p.category else '',
            p.supplier.name if p.supplier else '',
            p.warehouse.name if p.warehouse else ''
        ])
    
    # Lưu file
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(output, 
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    as_attachment=True,
                    download_name=f'inventory_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx')

@app.route('/api/low-stock')
def api_low_stock():
    products = Product.query.filter(Product.quantity <= Product.min_stock).all()
    return jsonify([{
        'id': p.id,
        'name': p.name,
        'sku': p.sku,
        'quantity': p.quantity,
        'min_stock': p.min_stock
    } for p in products])

@app.route('/suppliers')
def suppliers():
    suppliers = Supplier.query.all()
    return render_template('suppliers.html', suppliers=suppliers)

@app.route('/supplier/add', methods=['POST'])
def add_supplier():
    supplier = Supplier(
        name=request.form['name'],
        contact=request.form.get('contact'),
        email=request.form.get('email'),
        phone=request.form.get('phone')
    )
    db.session.add(supplier)
    db.session.commit()
    flash('Nhà cung cấp đã được thêm!', 'success')
    return redirect(url_for('suppliers'))

@app.route('/warehouses')
def warehouses():
    warehouses = Warehouse.query.all()
    return render_template('warehouses.html', warehouses=warehouses)

@app.route('/warehouse/add', methods=['POST'])
def add_warehouse():
    warehouse = Warehouse(
        name=request.form['name'],
        location=request.form.get('location')
    )
    db.session.add(warehouse)
    db.session.commit()
    flash('Kho đã được thêm!', 'success')
    return redirect(url_for('warehouses'))

# Initialize database
with app.app_context():
    db.create_all()
    
    # Add sample data if empty
    if Category.query.count() == 0:
        categories = [
            Category(name='Điện tử'),
            Category(name='Thực phẩm'),
            Category(name='Quần áo'),
            Category(name='Đồ gia dụng')
        ]
        db.session.add_all(categories)
        db.session.commit()

if __name__ == '__main__':
    app.run(port=5000)